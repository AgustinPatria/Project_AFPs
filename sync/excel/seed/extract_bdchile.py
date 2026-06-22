"""Extract BDChile master (Sec 05 Chilean stocks dimensional) from legacy Excel.

Source : Excels construccion pdf/<periodo>/20_Cartera_Acc_Chilenas.xlsm / BDChile
Output : validacion/bdchile.json

12 columns per row: nemo, ticker BBG, GICS (3 levels), Cuartil (1-4 or Other),
Sector interno, Company name, Group name, GICS name, Moneda Gral ID, Pais.

Read-only. Nothing is written to Supabase here — use load_bdchile.py for that.
"""
import argparse
import json
import sys
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parents[3]
EXCEL_BASE = ROOT / "Excels construccion pdf"
OUT = ROOT / "validacion" / "bdchile.json"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--periodo",
        required=True,
        help="YYYYMM subfolder under 'Excels construccion pdf/' (e.g. 202511)",
    )
    args = parser.parse_args()

    xlsm = EXCEL_BASE / args.periodo / "20_Cartera_Acc_Chilenas.xlsm"
    if not xlsm.exists():
        sys.exit(f"Excel not found: {xlsm}")
    print(f"Reading: {xlsm}")

    wb = openpyxl.load_workbook(xlsm, read_only=True, data_only=True)
    ws = wb["BDChile"]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        cols = row[:12]
        nemo = cols[0]
        if not isinstance(nemo, str) or not nemo.strip():
            continue
        rows.append(
            {
                "nemo": nemo.strip(),
                "ticker_bbg": cols[1].strip() if isinstance(cols[1], str) else None,
                "gics_8d": cols[2] if cols[2] is not None else None,
                "gics_2d": cols[3] if cols[3] is not None else None,
                "gics_4d": cols[4] if cols[4] is not None else None,
                "cuartil": str(cols[5]).strip() if cols[5] is not None else None,
                "sector_interno": cols[6].strip() if isinstance(cols[6], str) else None,
                "company_name": cols[7].strip() if isinstance(cols[7], str) else None,
                "group_name": cols[8].strip() if isinstance(cols[8], str) else None,
                "gics_name": cols[9].strip() if isinstance(cols[9], str) else None,
                "moneda_gral_id": cols[10] if cols[10] is not None else None,
                "pais": cols[11].strip() if isinstance(cols[11], str) else None,
            }
        )

    OUT.write_text(json.dumps(rows, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Extracted {len(rows)} rows -> {OUT}")

    # Quick stats
    from collections import Counter
    cuart = Counter(r["cuartil"] for r in rows)
    print("\nBy cuartil:")
    for k, v in sorted(cuart.items(), key=lambda x: (x[0] or "")):
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
