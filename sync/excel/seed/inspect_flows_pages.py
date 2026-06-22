"""READ-ONLY: layout y valores de las hojas Flows1/Flows2 (paginas 4-5 PDF Sec 07)
y ResumenFlows (12_Flows04) del set 202511."""
import sys
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parents[3] / "Excels construccion pdf" / "202511"


def dump(path, sheet, max_row, max_col):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    print(f"\n=== {path.name} [{sheet}] {ws.max_row}x{ws.max_column} ===")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True), 1):
        vals = []
        for v in row:
            if v is None:
                vals.append("")
            elif isinstance(v, float):
                vals.append(f"{v:,.1f}")
            else:
                vals.append(str(v)[:22])
        print(f"{i:>3} | " + " | ".join(vals))
    wb.close()


if __name__ == "__main__":
    dump(BASE / "13_Returns y Flows.xlsm", "Flows1", 45, 10)
    dump(BASE / "13_Returns y Flows.xlsm", "Flows2", 45, 10)
    dump(BASE / "12_Flows04.xlsm", "ResumenFlows", 15, 12)
