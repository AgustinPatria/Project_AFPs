"""Extract Sec 08 Top Net Inflows/Outflows from 19_Top_Purchases_Sales.xlsm.

Source : Excels construccion pdf/19_Top_Purchases_Sales.xlsm / Flows
Output : validacion/sec08_flows.json

Sheet layout (R = row):
  R5:  Title with date in parens (snapshot)
  R7:  MONTHLY
  R8:  NET INFLOWS (TOP 10)
  R9..R18: rk + fondo + amount (col H)
  R19: NET OUTFLOWS (TOP 10)
  R20..R29: rk + fondo + amount
  R31: YTD
  R32: NET INFLOWS
  R33..R42: rk + fondo + amount
  R43: NET OUTFLOWS
  R44..R53: rk + fondo + amount

Periodo del Excel actual: 31-Mar-2026.
"""
import json
import re
from pathlib import Path
from datetime import datetime, date
import openpyxl

XLSM = (
    Path(__file__).resolve().parents[3]
    / "Excels construccion pdf"
    / "19_Top_Purchases_Sales.xlsm"
)
OUT = Path(__file__).resolve().parents[3] / "validacion" / "sec08_flows.json"

# Section boundaries (1-indexed rows from the sheet)
SECTIONS = [
    ("MTD", "inflow", 9, 18),
    ("MTD", "outflow", 20, 29),
    ("YTD", "inflow", 33, 42),
    ("YTD", "outflow", 44, 53),
]


def main():
    wb = openpyxl.load_workbook(XLSM, read_only=False, data_only=True)
    ws = wb["Flows"]

    # Extract snapshot date from row 5 title: "Top Net Inflows and Outflows - Foreign Funds (31-Mar-2026)"
    title = None
    for c in ws[5]:
        if isinstance(c.value, str) and "(" in c.value:
            title = c.value
            break
    m = re.search(r"\((\d{1,2}-\w{3}-\d{4})\)", title or "")
    fecha = "unknown"
    if m:
        try:
            d = datetime.strptime(m.group(1), "%d-%b-%Y").date()
            fecha = d.isoformat()
        except ValueError:
            pass
    print(f"Snapshot fecha: {fecha}")

    rows = []
    for period_type, direction, r_start, r_end in SECTIONS:
        for r in range(r_start, r_end + 1):
            rk_cell = ws.cell(row=r, column=1).value
            fondo_cell = ws.cell(row=r, column=2).value
            amount_cell = ws.cell(row=r, column=8).value
            if not isinstance(rk_cell, (int, float)) or not fondo_cell:
                continue
            try:
                amount = float(amount_cell)
            except (TypeError, ValueError):
                continue
            rows.append({
                "fecha": fecha,
                "period_type": period_type,
                "direction": direction,
                "rk": int(rk_cell),
                "fondo": str(fondo_cell).strip(),
                "amount_usd_mm": amount,
            })

    OUT.write_text(json.dumps(rows, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Extracted {len(rows)} rows -> {OUT}")

    # Stats
    from collections import Counter
    by_section = Counter((r["period_type"], r["direction"]) for r in rows)
    for k, v in sorted(by_section.items()):
        print(f"  {k[0]} {k[1]}: {v}")


if __name__ == "__main__":
    main()
