"""Extract IPSA composition + Mkt_Cap + Pionero + MRV from legacy Excel.

Source : Excels construccion pdf/<periodo>/20_Cartera_Acc_Chilenas.xlsm
Outputs: validacion/ipsa.json, validacion/mkt_cap.json, validacion/benchmark_composition.json

The periodo (YYYYMM) is required and selects which monthly subfolder to read.
FechaReporte is read from the "Fecha" sheet (cell L16) and stored as ISO date
on every JSON row.
"""
import argparse
import json
import sys
from pathlib import Path
from datetime import datetime, date
import openpyxl

ROOT = Path(__file__).resolve().parents[3]
EXCEL_BASE = ROOT / "Excels construccion pdf"
OUTDIR = ROOT / "validacion"


def excel_path(periodo: str) -> Path:
    return EXCEL_BASE / periodo / "20_Cartera_Acc_Chilenas.xlsm"


def _to_str(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return str(v).strip() or None


def _to_float(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _to_date(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


def extract_periodo(wb) -> str:
    """Read FechaReporte from the Fecha sheet (cell L16 in the actual file)."""
    ws = wb["Fecha"]
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if any(c == "FechaReporte" for c in row if isinstance(c, str)):
            # The value is in the next column over.
            for i, c in enumerate(row):
                if c == "FechaReporte":
                    return str(row[i + 1]) if i + 1 < len(row) else "unknown"
    return "unknown"


def extract_ipsa(wb, fecha: str):
    ws = wb["IPSA"]
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        ticker = _to_str(row[0])
        if not ticker:
            continue
        rows.append({
            "fecha": fecha,
            "ticker_bbg": ticker,
            "weight": _to_float(row[1]),
            "gics_2d": _to_str(row[2]),
            "cuartil": _to_str(row[3]),
            "sector_interno": _to_str(row[4]) if len(row) > 4 else None,
        })
    return rows


def extract_mkt_cap(wb):
    ws = wb["Mkt_Cap"]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        fecha = _to_date(row[0])
        ticker = _to_str(row[1])
        mkt_cap = _to_float(row[2])
        if not fecha or not ticker or mkt_cap is None:
            continue
        rows.append({
            "fecha": fecha,
            "ticker_bbg": ticker,
            "mkt_cap_usd_mm": mkt_cap,
        })
    return rows


def extract_benchmark(wb, sheet_name: str, fecha: str, benchmark: str):
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        nemo = _to_str(row[0])
        if not nemo:
            continue
        weight = _to_float(row[1])
        if weight is None:
            continue
        rows.append({
            "fecha": fecha,
            "benchmark": benchmark,
            "nemo": nemo,
            "weight": weight,
            "gics_2d": _to_str(row[2]),
            "cuartil": _to_str(row[3]),
            "sector_interno": _to_str(row[4]) if len(row) > 4 else None,
        })
    return rows


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--periodo",
        required=True,
        help="YYYYMM subfolder under 'Excels construccion pdf/' (e.g. 202511)",
    )
    args = parser.parse_args()

    xlsm = excel_path(args.periodo)
    if not xlsm.exists():
        sys.exit(f"Excel not found: {xlsm}")
    print(f"Reading: {xlsm}")

    wb = openpyxl.load_workbook(xlsm, read_only=True, data_only=True)
    periodo_raw = extract_periodo(wb)
    print(f"Excel periodo: {periodo_raw}")

    # Convert "31-Mar-2026" -> "2026-03-31"
    try:
        periodo_dt = datetime.strptime(periodo_raw, "%d-%b-%Y").date()
    except ValueError:
        periodo_dt = date.today()
        print(f"  Warning: could not parse {periodo_raw!r}, using today")
    fecha = periodo_dt.isoformat()
    print(f"  Using fecha = {fecha}")

    ipsa = extract_ipsa(wb, fecha)
    (OUTDIR / "ipsa.json").write_text(
        json.dumps(ipsa, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    print(f"IPSA: {len(ipsa)} rows -> ipsa.json")

    mkt = extract_mkt_cap(wb)
    (OUTDIR / "mkt_cap.json").write_text(
        json.dumps(mkt, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    fechas = sorted({r["fecha"] for r in mkt})
    print(f"Mkt_Cap: {len(mkt)} rows across {len(fechas)} fechas -> mkt_cap.json")

    pio = extract_benchmark(wb, "Pionero", fecha, "Pionero")
    mrv = extract_benchmark(wb, "MRV", fecha, "MRV")
    bench = pio + mrv
    (OUTDIR / "benchmark_composition.json").write_text(
        json.dumps(bench, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    print(f"Benchmarks: Pionero {len(pio)} + MRV {len(mrv)} = {len(bench)} rows -> benchmark_composition.json")


if __name__ == "__main__":
    main()
