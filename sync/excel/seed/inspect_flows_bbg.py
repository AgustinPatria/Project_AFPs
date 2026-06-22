"""Diagnostico READ-ONLY del pipeline de retornos legacy (Sec 07 pags 4-5).
Inspecciona 09_Flows01.xlsm (tickers Bloomberg) y 11_Flows03.xlsm (Rentab/Return).
"""
import sys
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parents[3] / "Excels construccion pdf" / "202511"


def dump(path, sheet, max_row=10, max_col=12, data_only=True):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=data_only)
    ws = wb[sheet]
    print(f"\n=== {path.name} [{sheet}] {ws.max_row}x{ws.max_column} (data_only={data_only}) ===")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True), 1):
        print(f"{i:>3} | " + " | ".join("" if v is None else str(v)[:24] for v in row))
    wb.close()


if __name__ == "__main__":
    f09 = BASE / "09_Flows01.xlsm"
    f11 = BASE / "11_Flows03.xlsm"
    dump(f09, "BaseDatos_Tickers", max_row=12, max_col=8)
    dump(f09, "BBG", max_row=10, max_col=8)
    dump(f09, "BBG", max_row=10, max_col=8, data_only=False)  # ver formulas BDH/BDP
    dump(f09, "Value", max_row=8, max_col=8)
    dump(f11, "Rentab", max_row=10, max_col=10)
    dump(f11, "Return", max_row=10, max_col=10)
    dump(f11, "Change", max_row=8, max_col=10)
    dump(f11, "Flows", max_row=8, max_col=10)
