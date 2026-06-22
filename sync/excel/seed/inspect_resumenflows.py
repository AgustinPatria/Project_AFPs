"""READ-ONLY: (1) tablas de lookup N10:R22 de Flows1; (2) header + formulas de
ResumenFlows en 12_Flows04 para entender las columnas Return/Flows por ventana."""
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parents[3] / "Excels construccion pdf" / "202511"

print("=== Flows1!N5:R23 (lookup ventana -> columnas) ===")
wb = openpyxl.load_workbook(BASE / "13_Returns y Flows.xlsm", read_only=True, data_only=True)
ws = wb["Flows1"]
for i, row in enumerate(ws.iter_rows(min_row=5, max_row=23, min_col=14, max_col=18), 5):
    vals = ["" if getattr(c, "value", None) is None else str(c.value)[:18] for c in row]
    if any(vals):
        print(f" r{i} | " + " | ".join(vals))
wb.close()

print("\n=== 12_Flows04 [ResumenFlows] headers (filas 4-6, cols N..AC) ===")
wb = openpyxl.load_workbook(BASE / "12_Flows04.xlsm", read_only=True, data_only=True)
ws = wb["ResumenFlows"]
for i, row in enumerate(ws.iter_rows(min_row=3, max_row=6, min_col=14, max_col=29), 3):
    vals = ["" if getattr(c, "value", None) is None else str(c.value)[:16] for c in row]
    if any(vals):
        print(f" r{i} | " + " | ".join(vals))
wb.close()

print("\n=== 12_Flows04 [ResumenFlows] formulas fila 7 (cols N..AC) ===")
wb = openpyxl.load_workbook(BASE / "12_Flows04.xlsm", read_only=True, data_only=False)
ws = wb["ResumenFlows"]
for row in ws.iter_rows(min_row=7, max_row=7, min_col=14, max_col=29):
    for j, c in enumerate(row, 14):
        v = getattr(c, "value", None)
        if v is not None:
            print(f"  {get_column_letter(j)}7: {str(v)[:180]}")
wb.close()
