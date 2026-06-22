"""READ-ONLY: formula completa de una celda de datos de Change/Return/Flows."""
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

XLSM = (
    Path(__file__).resolve().parents[3]
    / "Excels construccion pdf" / "202511" / "11_Flows03.xlsm"
)

wb = openpyxl.load_workbook(XLSM, read_only=True, data_only=False)
for sheet in ("Change", "Return", "Flows"):
    ws = wb[sheet]
    print(f"\n=== [{sheet}] ===")
    # encabezados de columnas de datos (filas 4-6, cols 14-18) y formula completa fila 7
    for row in ws.iter_rows(min_row=4, max_row=7, min_col=14, max_col=17):
        for c in row:
            v = getattr(c, "value", None)
            if v is not None:
                col = get_column_letter(getattr(c, "column", 0)) if hasattr(c, "column") else "?"
                r = getattr(c, "row", "?")
                print(f"  {col}{r}: {str(v)}")
wb.close()
