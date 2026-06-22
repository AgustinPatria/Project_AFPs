"""READ-ONLY: verifica el bug de truncamiento del legacy — las matrices
Change/Return/Flows de 11_Flows03 cubren filas 7..2000, pero Output_25sd tiene
mas instrumentos. Mide el cambio Oct->Nov de las filas excluidas (>2000), por
asset class y region, que deberia explicar el residuo del panel del PDF."""
from collections import defaultdict
from pathlib import Path

import openpyxl

XLSM = (
    Path(__file__).resolve().parents[3]
    / "Excels construccion pdf" / "202511" / "11_Flows03.xlsm"
)

wb = openpyxl.load_workbook(XLSM, read_only=True, data_only=True)
ws = wb["Output_25sd"]
print(f"Output_25sd en 11_Flows03: {ws.max_row} filas x {ws.max_column} cols")

# mapeo confirmado por la lookup table de Flows1: GZ (col 208) = 2025-11-30, GY (207) = 2025-10-31
col_nov, col_oct = 208, 207

# sumar cambio Oct->Nov de filas excluidas (>2000) por (asset_class, region)
exc = defaultdict(float)
n_exc = 0
for i, row in enumerate(ws.iter_rows(min_row=2001, max_col=max(col_nov, col_oct), values_only=True), 2001):
    ident = row[0]
    if not isinstance(ident, str) or not ident.strip():
        continue
    ac, tipo, region = row[6], row[7], row[9]
    nov = row[col_nov - 1] if isinstance(row[col_nov - 1], (int, float)) else 0
    oct_ = row[col_oct - 1] if isinstance(row[col_oct - 1], (int, float)) else 0
    if tipo == "[Direct Investment]" or ac not in ("Equity", "Fixed Income", "Private Equity"):
        continue
    n_exc += 1
    exc[(ac, region)] += nov - oct_
wb.close()

print(f"\nInstrumentos en filas >2000 (scope flows): {n_exc}")
print(f"{'asset_class':<16} {'region':<22} {'cambio Oct->Nov':>15}")
for (ac, reg), v in sorted(exc.items(), key=lambda x: -abs(x[1])):
    if abs(v) > 0.5:
        print(f"{ac:<16} {str(reg):<22} {v:>15,.1f}")
print(f"{'TOTAL':<39} {sum(exc.values()):>15,.1f}")
