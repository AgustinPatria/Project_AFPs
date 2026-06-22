"""READ-ONLY: formulas exactas de las matrices Change/Return/Flows de 11_Flows03.xlsm
para reverse-engineer la metodologia del split del PDF."""
from pathlib import Path

import openpyxl

XLSM = (
    Path(__file__).resolve().parents[3]
    / "Excels construccion pdf" / "202511" / "11_Flows03.xlsm"
)

wb = openpyxl.load_workbook(XLSM, read_only=True, data_only=False)
for sheet in ("Change", "Return", "Flows"):
    ws = wb[sheet]
    print(f"\n=== [{sheet}] {ws.max_row}x{ws.max_column} ===")
    # fila 6 = header (ID, Bloomberg ID, ..., fechas). Mostrar header cols 1-20 y
    # algunas celdas de datos (filas 7-9, columnas 13-18) con sus formulas.
    for i, row in enumerate(
        ws.iter_rows(min_row=5, max_row=9, min_col=1, max_col=20), 5
    ):
        line = []
        for c in row:
            v = getattr(c, "value", None)
            if v is None:
                line.append("")
            else:
                line.append(str(v)[:60])
        print(f" r{i} | " + " | ".join(line))
wb.close()
