"""READ-ONLY: formulas de las celdas del panel Flows1 (13_Returns y Flows.xlsm)
para ver de donde salen las columnas Return/Flows YTD y Monthly."""
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

XLSM = (
    Path(__file__).resolve().parents[3]
    / "Excels construccion pdf" / "202511" / "13_Returns y Flows.xlsm"
)

wb = openpyxl.load_workbook(XLSM, read_only=True, data_only=False)
ws = wb["Flows1"]
# fila 11 = GEM subtotal, fila 14 = GEM Mixed (dato), cols C..J
for i, row in enumerate(ws.iter_rows(min_row=10, max_row=14, min_col=3, max_col=10), 10):
    for j, c in enumerate(row, 3):
        v = getattr(c, "value", None)
        if v is not None and isinstance(v, str) and v.startswith("="):
            print(f"{get_column_letter(j)}{i}: {v}")
    print("---")
wb.close()
