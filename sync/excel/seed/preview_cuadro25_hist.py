"""Preview del backfill historico del Cuadro 25 (Sec 07 Foreign).
Lee las hojas mensuales de 04_tabla_sin_desfase_25.xlsm y muestra estructura,
secciones, conteos y totales por mes. NO escribe nada en ninguna base.

Uso:
  python preview_cuadro25_hist.py peek 202511    # estructura + secciones + tail
  python preview_cuadro25_hist.py sweep          # todos los meses: filas y USD total
"""
import sys
from pathlib import Path

import openpyxl

XLSM = (
    Path(__file__).resolve().parents[3]
    / "Excels construccion pdf"
    / "202511"
    / "04_tabla_sin_desfase_25.xlsm"
)

HEADER_ROW = 5  # NEMOTECNICO | INSTITUCION | FONDO A..E | TOTAL | %


def parse_sheet(ws):
    """Devuelve (data_rows, secciones) de una hoja mensual del cuadro 25.
    data_row = (seccion, nemo, institucion, total_usd_mm)."""
    rows, secciones = [], []
    seccion = None
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, max_col=9, values_only=True):
        ident = row[0]
        if ident is None or (isinstance(ident, str) and not ident.strip()):
            continue
        ident = str(ident).strip()
        total = row[7]
        institucion = row[1]
        es_dato = isinstance(total, (int, float)) and institucion not in (None, "")
        if not es_dato:
            # fila de seccion (EMPRESAS EXTRANJERAS, etc.) o de cierre
            seccion = ident
            secciones.append(ident)
            continue
        rows.append((seccion, ident, str(institucion).strip(), float(total)))
    return rows, secciones


def peek(sheet_name: str):
    wb = openpyxl.load_workbook(XLSM, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows, secciones = parse_sheet(ws)
    print(f"=== hoja {sheet_name}: {ws.max_row} filas excel, {len(rows)} filas de datos ===")
    print("\nSecciones encontradas (en orden):")
    for s in secciones:
        n = sum(1 for r in rows if r[0] == s)
        usd = sum(r[3] for r in rows if r[0] == s)
        print(f"  {s:<50} {n:>5} filas  {usd:>12,.0f} USD mm")
    print(f"\nTOTAL datos: {sum(r[3] for r in rows):>12,.0f} USD mm")
    print("\nPrimeras 3 filas de datos:")
    for r in rows[:3]:
        print("  ", r)
    print("\nUltimas 5 filas de datos:")
    for r in rows[-5:]:
        print("  ", r)
    wb.close()


def sweep():
    wb = openpyxl.load_workbook(XLSM, read_only=True, data_only=True)
    meses = sorted(n for n in wb.sheetnames if n.isdigit() and len(n) == 6)
    print(f"{len(meses)} hojas mensuales: {meses[0]} .. {meses[-1]}\n")
    print(f"{'mes':>8} {'filas':>6} {'usd_mm':>12}")
    for m in meses:
        rows, _ = parse_sheet(wb[m])
        print(f"{m:>8} {len(rows):>6} {sum(r[3] for r in rows):>12,.0f}")
    wb.close()


def export(desde: str, hasta: str):
    """Escribe a CSV las filas exactas que se cargarian (sin tocar ninguna BD)."""
    import csv

    out = Path(__file__).resolve().parents[3] / "validacion" / "foreign_cuadro25_hist_preview.csv"
    wb = openpyxl.load_workbook(XLSM, read_only=True, data_only=True)
    meses = sorted(n for n in wb.sheetnames if n.isdigit() and len(n) == 6 and desde <= n <= hasta)
    n_total = 0
    with out.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["periodo", "seccion", "nemo_isin", "institucion", "monto_usd_mm"])
        for m in meses:
            rows, _ = parse_sheet(wb[m])
            for seccion, nemo, institucion, total in rows:
                w.writerow([f"{m[:4]}-{m[4:]}", seccion, nemo, institucion, total])
            n_total += len(rows)
    wb.close()
    print(f"{n_total} filas ({meses[0]}..{meses[-1]}, {len(meses)} meses) -> {out}")


if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else "peek"
    if cmd == "sweep":
        sweep()
    elif cmd == "export":
        export(sys.argv[2] if len(sys.argv) > 2 else "200909",
               sys.argv[3] if len(sys.argv) > 3 else "202412")
    else:
        peek(sys.argv[2] if len(sys.argv) > 2 else "202511")
