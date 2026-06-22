"""Extract ALL 2,182 foreign-classification rows from the legacy Excel master.
Saves to validacion/foreign_overlay.json — NOTHING is written to Supabase here.
"""
import json
from pathlib import Path
import openpyxl

XLSM = (
    Path(__file__).resolve().parents[1]
    / "Excels construccion pdf"
    / "04_tabla_sin_desfase_25.xlsm"
)


def main():
    wb = openpyxl.load_workbook(XLSM, read_only=True, data_only=True)
    ws = wb["Output_25sd"]
    rows = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        ident = row[0]
        if not isinstance(ident, str) or not ident.strip():
            continue
        rows.append(
            {
                "identificador": ident.strip(),
                "alt_id": row[1] if isinstance(row[1], str) else None,
                "family": row[2],
                "manager": row[3],
                "fondo": row[4],
                "asset_class": row[6],
                "fund_type": row[7],
                "fund_style": row[8],
                "region": row[9],
                "country": row[10],
                "category": row[11],
                "currency": row[12],
            }
        )

    out_json = Path(__file__).resolve().parents[1] / "validacion" / "foreign_overlay.json"
    out_json.write_text(json.dumps(rows, indent=2, default=str), encoding="utf-8")
    print(f"Extracted {len(rows)} rows -> {out_json}")

    by_ac = {}
    for r in rows:
        by_ac[r["asset_class"]] = by_ac.get(r["asset_class"], 0) + 1
    print("\nBy asset_class:")
    for k, v in sorted(by_ac.items(), key=lambda x: -x[1]):
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
