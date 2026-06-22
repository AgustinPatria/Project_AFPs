"""Extract the manual Direct Investment overlay from the legacy Excel and
emit a SQL INSERT script. One-shot extractor — run once, commit the SQL
(or just apply via apply_migration).
"""
import json
from pathlib import Path
import openpyxl

XLSM = (
    Path(__file__).resolve().parents[1]
    / "Excels construccion pdf"
    / "04_tabla_sin_desfase_25.xlsm"
)


def main():
    wb = openpyxl.load_workbook(XLSM, read_only=True, data_only=True)
    ws = wb["Output_25sd"]
    rows = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        # Columns (0-indexed): 0=identificador, 1=alt id, 2=family, 3=manager,
        # 4=fondo, 5=NO flag, 6=asset_class, 7=fund_type, 8=fund_style,
        # 9=region, 10=country, 11=di_category, 12=currency
        if row[7] != "[Direct Investment]" and row[8] != "[Direct Investment]":
            continue
        identificador = (row[0] or "").strip() if isinstance(row[0], str) else row[0]
        if not identificador:
            continue
        rows.append(
            {
                "identificador": identificador,
                "emisor_norm": (row[4] or "").strip() if isinstance(row[4], str) else None,
                "asset_class": row[6],
                "region": row[9],
                "country": row[10],
                "di_category": row[11] if row[11] in ("Sovereign", "Bank", "Corporate") else None,
                "currency": row[12] if isinstance(row[12], str) and row[12] != "[Direct Investment]" else None,
            }
        )

    print(f"Extracted {len(rows)} Direct Investment rows")

    out_json = Path(__file__).resolve().parents[1] / "validacion" / "di_overlay.json"
    out_json.write_text(json.dumps(rows, indent=2, default=str), encoding="utf-8")
    print(f"Wrote {out_json}")

    cats = {}
    for r in rows:
        k = (r["asset_class"], r["di_category"])
        cats[k] = cats.get(k, 0) + 1
    print("\nBreakdown asset_class × di_category:")
    for k, v in sorted(cats.items()):
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
